import os
from openpyxl import load_workbook
from utilities.config_reader import ConfigReader


class ExcelReader:
    """
    Class Name    : ExcelReader
    Author        : Saptarshi
    Description   : Utility class responsible for reading test data from Excel files using openpyxl library
    Return Type   : Object
    Parameters    : None
    """

    def __init__(self):
        try:
            # Initialize ConfigReader to fetch configuration values
            config_reader = ConfigReader()

            # Read relative Excel file path from config.properties
            excel_file_relative_path = config_reader.get_config_value(
                "PATH",
                "excel_path"
            )

            # Validate Excel path availability
            if not excel_file_relative_path:
                raise ValueError(
                    "Excel file path not found in config.properties"
                )

            # Resolve project root directory dynamically
            project_root_directory = os.path.dirname(
                os.path.dirname(os.path.abspath(__file__))
            )

            # Construct absolute path of Excel file
            excel_file_absolute_path = os.path.join(
                project_root_directory,
                excel_file_relative_path
            )

            # Verify Excel file existence
            if not os.path.exists(excel_file_absolute_path):
                raise FileNotFoundError(
                    f"Excel file not found at: {excel_file_absolute_path}"
                )

            # Load Excel workbook into memory
            self.excel_workbook = load_workbook(
                excel_file_absolute_path
            )

        except Exception as exception:
            # Raise meaningful exception if workbook loading fails
            raise Exception(
                f"Failed to load Excel workbook: {exception}"
            )

    """
    Method Name   : get_cell_value
    Author        : Saptarshi
    Description   : Fetches data from a specific cell in an Excel sheet using row and column numbers
    Return Type   : Any
    Parameters    : sheet_name(str), row_number(int), column_number(int)
    """

    def get_cell_value(self, sheet_name, row_number, column_number):
        try:
            # Validate whether the sheet exists in the workbook
            if sheet_name not in self.excel_workbook.sheetnames:
                raise ValueError(
                    f"Sheet '{sheet_name}' does not exist in Excel file"
                )

            # Access the required sheet
            excel_sheet = self.excel_workbook[sheet_name]

            # Fetch and return cell value
            return excel_sheet.cell(
                row=row_number,
                column=column_number
            ).value

        except Exception as exception:
            # Raise exception with detailed context
            raise Exception(
                f"Failed to read data from sheet '{sheet_name}', "
                f"row {row_number}, column {column_number}: {exception}"
            )
 